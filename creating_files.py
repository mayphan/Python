import mysql.connector
import os,sys,logging,json
import pandas as pd
import openpyxl,time
from openpyxl import load_workbook
# from list_col_viri import viri_label_list
from datetime import datetime,timedelta
from create_connection import create_connectn


# def writecol_to_exl(root,strt_time,sql_col):
#     # print(root)
#     xf = openpyxl.load_workbook(filename=root)
#     sheet1 = xf['Sheet']
#     sheet1.cell(row=1,column=1,value="TIMESTAMP")
#     xf.save(root)
#     ro,col=1,2    
   
    # for val in range(4,len(sql_col)):                        
    #     #appending       
    #     sheet1.cell(row=ro,column=col,value=sql_col[val])                
    #     xf.save(root)
    #     col+=1
    #     xf.save(root)
    # time_tkn=time.time()-strt_time
    # print("columns created ,took seconds:",time_tkn)
           
        


def create_file(prev_date):
    try:
        
        strt_time=time.time()
        curr_date=datetime.now().strftime('%Y-%m-%d')
        # print(curr_date)
        # prev_date= (datetime.now() + timedelta(days=-1)).strftime('%Y-%m-%d') 
        # print(prev_date)
        # db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data_testing` WHERE `date_` LIKE %(p)s ORDER BY date_ ASC,time_ ASC" 
        db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data` WHERE `date_` LIKE %(p)s ORDER BY date_ ASC,time_ ASC"
        sql_df,veh_df,veh_logr,veh_nam,df_col=create_connectn(curr_date,prev_date,db_qry,0,0)
        # print(sql_col)
        first_veh=sql_df['vehicle_name'][0]
        # print(first_veh)
        veh_name_list,empt_list=[],[]
        for row_ in range(0,len(sql_df)): 
            empt_list.append(sql_df["vehicle_name"][row_])
        for veh in empt_list:
            if veh in veh_name_list:
                # print(veh)
                pass
            else:
                # print(veh)
                veh_name_list.append(veh)
        # print(veh_name_list)
        RRD_path_list=[]
        for vehcl in veh_name_list:        
            veh_name=veh_nam[vehcl]
            format_name_2="{}_{}_{}.xlsx".format(prev_date,veh_name,veh_logr[vehcl])
            # print(format_name_2)
            RRD_file=os.path.join("E:\\","Company","Shared_Data","DataProcessing","FleetData",veh_name,"RefinedRawData",format_name_2) 
            # print(RRD_file.split("FleetData\\")[1].split("\\")[0])
            RRD_path_list.append(RRD_file)  
            # print(RRD_file)      
            wb = openpyxl.Workbook()
            wb.save(RRD_file)
            # writecol_to_exl(RRD_file,strt_time,sql_col)
            print("\n----------------------------Excel created for vehicle: {}--------------------------------------------------------\n".format(veh_name))
            
        # print(RRD_path_list,veh_name_list,RRD_path_list,veh_name_list,sql_df,veh_df,veh_logr,veh_nam,df_col)  
        return RRD_path_list,veh_name_list,sql_df,veh_df,veh_logr,veh_nam,df_col
    except:
        print("Raw Data not observed in DB for date: ",prev_date)
        
        
    



         