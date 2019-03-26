import creating_excl
import pandas as pd
import openpyxl

#calling function from imported excl file
excl.createExcel()


def write_in_existexcl(list1,ind):
	
	file1='Contacts.xlsx'
	file2=openpyxl.load_workbook(filename=file1)
	contct_sheet=file2['Contacts']
	sz=len(ind)+1
	sz1=len(list1)	
	k,m,x=0,0,0
	j,l=2,1
	while(x<4):
		for i in range(1,sz):
		#	j,l=2,1
			contct_sheet.cell(row=j+i,column=l).value=list1[m][ind[k]]
			k=k+1
		j=2
		l=l+1
		m=m+1
		x=x+1
		k=0
	j,l,m,k=2,7,4,0			
	while(x<sz1):
                for i in range(1,sz):
                #       j,l=2,1
                        contct_sheet.cell(row=j+i,column=l).value=list1[m][ind[k]]
                        k=k+1
                j=2
                l=l+1
                m=m+1
                x=x+1
                k=0
	
	file2.save('Contacts.xlsx')


def check_info(df1,cs):
	#if df1['columns'] not in cs:
	#	print("missing columns")
	#else:
	#selecting contacts on basis of any of one available information: home or work adress/email/phone
	df2=df1[((df1['home_address_1'].notnull() | df1['home_address_2'].notnull()) & df1['home_city'].notnull() & df1['home_state'].notnull()
	& df1['home_postal_code'].notnull() & df1['home_country'].notnull()) | ((df1['work_address_1'].notnull() | 
	df1['work_address_2'].notnull()) & df1['work_city'].notnull() & df1['work_state'].notnull() & 
	df1['work_postal_code'].notnull() & df1['work_country'].notnull()) |(df1['home_email'].notnull()|df1['main_email']|
	df1['other_email'].notnull() | df1['work_email'].notnull()) | ( df1['fax_phone'].notnull() | df1['home_phone'].notnull()|
	df1['mobile_phone'].notnull())]
	#print(df2['last_name'])
	#for i in range(0,147):	
	#	if df2[df2.columns[i]] not in cs:
	#		print("missing columns")
	#	else:
	
	#created list for inputing data from selected columns into excel sheet
	list1=[df2['last_name'],df2['first_name'],df2['company'],df2['title'],df2['work_phone'],df2['work_email'],
	df2['work_address_1'],df2['work_city'],df2['work_state'],df2['work_postal_code'],df2['home_address_1'],df2['home_city'],
	df2['home_state'],df2['home_postal_code'],df2['home_phone'],df2['home_email']]	
	#print(list1)
	#counting index of dataframe	
	ind=df2.index.tolist()

	#calling function
	write_in_existexcl(list1,ind)


def reading_csv():
	cs=pd.read_csv("exp1.csv", index_col=False)
	#cs=pd.read_csv("srch.csv")
	#selecting dataonly if both first and last name are present
	df1=cs[cs['first_name'].notnull() & cs['last_name'].notnull()]

	#calling function
	check_info(df1,cs)
	




if __name__=="__main__":
	
	reading_csv()



