import pandas as pd
import mysql.connector
import math,os,sys
from os import walk
from functools import reduce
import openpyxl
from openpyxl import load_workbook


def write_val(stmp,value_rcvd,indx_1,indx_2,root):
    print(root)
    df=pd.read_excel(root)
    col_=df.columns.get_loc(stmp)
    print(col_)
    ro=indx_1+1
    for range in (indx_1,indx_2):    
        xf2 = openpyxl.load_workbook(filename=root)
        sheet = xf2['Sheet1']       
              
        sheet.cell(row=ro,column=col_,value=value_rcvd)
        # print(raw_data_list.filelst_last[i][0].split("T")[0])
        xf2.save(root)
        ro+=1


def chk_zeroTH_indx(val_present):
    indx_list=[]
    for idx,vals in val_present.items():
        # print(idx)
        indx_list.append(idx)
    for ele in indx_list:
        if ele==0:
            indx_var='first_indx_present'
            
            return indx_list,indx_var
        else:
            pass
    indx_var='first_indx_absent'
    print(indx_list)
    return indx_list,indx_var


def find_if_data(val_chk):
    val_dict={}
    for data in val_chk:
        if data=='nan':
            pass
        else:
            idx=val_chk.index(data)
            val_dict[idx]=data
            # key_dict=val_dict.keys()
            
    return val_dict
            

def fill_values(val,stmp,root,dir_path):
    df=pd.read_excel(file_name)
    col=df.columns.get_loc(stmp)
    print(col)
    for indx in range(len(df)):
        xf = openpyxl.load_workbook(filename=root)
        sheet1 = xf['Sheet']
        sheet1.cell(row=indx+1,column=col,value=val)
        xf.save(root)
    print("{} column is written".format(stmp))


def create_connectn(prev_date,db_qry):
    print('inside connection')
    
    cnctn = mysql.connector.connect(host='tpsan1srv01',
        user='root',
        passwd='FifthRiver7#',
        db='telematics')
    cursor = cnctn.cursor() 

    cursor.execute(db_qry)
    default_df=pd.DataFrame(cursor.fetchall())
    default_df.columns = cursor.column_names
    # print(default_df)
    # print(default_df.columns)
    cnctn.commit()  
    cursor.close()
    cnctn.close() 
   
    return default_df,default_df.columns

def get_valid_val(df_for_col,idx,i):
    # print("idx:",idx)  
        
    try:
        if idx>=0:
            last_col_val=df_for_col[idx]
            # print(last_col_val)            
            
            if last_col_val=='nan':
                i+=1        
                indx=len(df_for_col)-i
                return get_valid_val(df_for_col,indx,i)
            
            else:
                # print(df_for_col)
                return last_col_val
        else:
            last_col_val='notfound'
            # print(last_col_val)
            return last_col_val


    except:        
        # print("excep")
        last_col_val='notfound'
        # print(last_col_val)
        return last_col_val
  

def check_old_dt(fil_dir):
    old_dts=[]
    for val in fil_dir:
        val=val.split('_')[0]
        if val not in old_dts:
            old_dts.append(val)
        else:
            pass
    print(old_dts)
    return old_dts


def find_last_val(val_chk,dir_path,prev_date,stmp):
    
    # print(val_chk)
    dict_col={}
    fil_dir=sorted(os.listdir(dir_path),reverse=True)
    # print(fil_dir)
    # print('-----------------------------------------------------------------------------------------------------------------------------------')
    # print(fil_dir)
    # print('-----------------------------------------------------------------------------------------------------------------------------------')
    # print(fil_dir)
    old_dt_list=check_old_dt(fil_dir)
    # try:
    for i in range(0,len(fil_dir)):
        try: 
        # for i in range(0,len(fil_dir)):
            # print(val_chk)
            if val_chk[i]=='nan':
                if prev_date>old_dt_list[i]:
                    # opn_file=[os.path.join(RRD_DIR, filename) for filename in os.listdir(RRD_DIR) if filename.startswith("FC_EDD1_RefinedRawData")]
                    if fil_dir[i].startswith(old_dt_list[i]):
                        path_=os.path.join(dir_path,fil_dir[i])
                        print(path_)                       
                        df=pd.read_excel(path_)
                                           
                       
                        if len(df)<1:
                            pass
                        else:    
                            df_for_col=list(df[stmp])
                            df_for_col=find_NAN_int(df_for_col)                            
                            last_col_val=get_valid_val(df_for_col,len(df_for_col)-1,1) 
                        
                        
                        if last_col_val=='notfound':
                            pass
                        else:
                            dict_col[stmp]=last_col_val
                            # print(dict_col)
                            return dict_col
                    else:
                        pass              
        except:
            
            try:
                print("empty column for date:")
                dict_col[stmp]= 'notfound'
                # return dict_col 
            
            except:
                pass        

    # updated_val=find_last_val(val_chk,dir_path,prev_date,stmp)
        


def find_NAN_int(list_1):  
    # print(list_1)
    li_1=[]
    try:
        for i in range(0,len(list_1)):        
            # print(list_1[i])
            try:
                if math.isnan(list_1[i]):
                    # print("NAN")
                    li_1.append('nan')
                elif type(list_1[i])==int:
                    li_1.append(list_1[i])
                elif type(list_1[i])==float:    
                    li_1.append(list_1[i]) 
                            
                elif type(list_1[i])=='numpy.float64':    
                    li_1.append(list_1[i])     
                # elif math.isnan(list_1[i]):
                #     li_1.append('nan') 
                            
                else:
                    print('no val observed:',type(list_1[i]))
                    li_1.append('no val observed')
            except:
                try:
                    li_1.append('nan')
                except:
                    print('some NAN function err')
        # print(li_1)
        return li_1
               
    except: 
        pass
        # li_1.insert(i,'exception observed')
    
def chk_prev_files(val_chk,dir_path,prev_date,stmp):
    try:
        updated_val=find_last_val(val_chk,dir_path,prev_date,stmp)
        print('updated value:',updated_val)

        print('--------------------------------------------------------------------------------------------------------------------------------------------------')
        # try:
        
        if  updated_val=='None' or updated_val=='notfound'  :
            print("The value is:",updated_val)
            pass
            
            # db_qry="SELECT * FROM telematics.default_processing_values"
        #     dflt_df,dflt_col=create_connectn(prev_date,db_qry)
        # print(dflt_df)
        else:
            print("value is:",updated_val)
            return updated_val
        # except:
        #     pass
    except:
        print("Value is:",updated_val)
        # return updated_val
        pass
     
    #     print("Check for default values in DB")
    #     print('--------------------------------------------------------------------------------------------------------------------------------------------------')
    #     db_qry="SELECT * FROM telematics.default_processing_values"
    #     dflt_df,dflt_col=create_connectn(prev_date,db_qry)
    #     print(dflt_df[stmp][0])
        # print(dflt_col[0])
        # dict_dflt[stmp]=dflt_df[stmp][0]
        # # print(dict_dflt)
        # def_list.append(dict_dflt)
        # stmp_list.append(stmp)


def last_label_val(file_,dir_path,prev_date):
    print(file_)
    df=pd.read_excel(file_,dtype=object)
    if len(df)>1:
        # col_list=df.columns.values
        col='accelpedalposition'.upper()
        # print(col)
        col_list=[col]
        # print(df_col)
        # df=read_exl["TIMESTAMP"]
        # print(df)
        def_list,stmp_list=[],[]
        dict_dflt={}
        try:
            for stmp in col_list:
                # print(stmp)
                if stmp=="TIMESTAMP":
                    pass
                else:
                    print('--------------------------------------------------------------------------------------------------------------------------------------------------')
                    # try:
                        
                    # print("length:", len(df[stmp]))
                    # check for NAN:
                    val_chk=find_NAN_int(df[stmp])
                    print(val_chk)
                    # check for index:
                    val_present=find_if_data(val_chk)
                    # print(len(val_present))
                    if len(val_present)>0:                    
                        indx_lst,zeroTh_indx=chk_zeroTH_indx(val_present)        
                        # print(indx_lst)            
                        if zeroTh_indx=='first_indx_present':                
                            # nxt_indx=chk_othr_val()
                            print("passing")
                            pass
                        else:
                            print("here")
                            value_rcvd=chk_prev_files(val_chk,dir_path,prev_date,stmp)  
                            print('RC:',value_rcvd[stmp])
                            #write value at zeroth
                            # write_val(stmp,value_rcvd[stmp],1,indx_lst,file_)                        

                    else:
                        # val_chk=find_NAN_int(df[stmp])                        
                        chk_prev_files(val_chk,dir_path,prev_date,stmp)
        except:
            try:
                print("Dataframe is empty")
            except:
                pass
        # updated_val=find_last_val(val_chk,dir_path,prev_date,stmp)
        # fill_values(dflt_df[stmp][0],stmp,file_name,dir_path)
    else:
        print("File is empty for vehicle: on date:")



sys.setrecursionlimit(100000)    
date_='2019-08-20'
vehicle_name="Peterbilt_220_EV_TNMC_UQM_EAxle_37"
# file_name="E:/Company/Shared_Data/DataProcessing/FleetData/Peterbilt_220_EV_TNMC_UQM_EAxle_37/RefinedRawData/2019-08-17_Peterbilt_220_EV_TNMC_UQM_EAxle_37_30475.xlsx"
file_name='C:/Users/mayuri/Documents/FleetData_1/210_1_KEN_T680_CPSTN_290915/RefinedRawData/2019-08-20_Peterbilt_220_EV_TNMC_UQM_EAxle_37_30475.xlsx'
# dir_path=os.path.join("E:\\","Company","Shared_Data","DataProcessing","FleetData",vehicle_name,"RefinedRawData")
dir_path='C:/Users/mayuri/Documents/FleetData_1/210_1_KEN_T680_CPSTN_290915/RefinedRawData'
    
# print(dir_path)
last_label_val(file_name,dir_path,date_)