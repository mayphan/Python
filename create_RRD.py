
# --------------------------depricated----------------

import mysql.connector
import logging
import os,sys,logging,json
from datetime import datetime,timedelta
import pandas as pd
import openpyxl,time
from openpyxl import load_workbook
from list_col_viri import viri_label_list
from functools import reduce

from creating_files import *
# from creating_files import RRD_path_list,veh_name_list
# from last_label_value import last_label_val



def RRDfile(veh_,veh_nam):
    for ele in range(0,len(RRD_path_list)):
        # print(veh_nam)
        if veh_nam[veh_]==RRD_path_list[0].split("FleetData\\")[1].split("\\")[0]:
            print(RRD_path_list[ele])
            return RRD_path_list[ele]
        else:
            pass
    

# def write_to_csv(df_data,df_veh,ro,row_,veh_,veh_nam):
def write_to_csv(df_data,ro,veh_,veh_nam):
    # veh_nam=
    root=RRDfile(veh_,veh_nam) 
    # print(root)   
    df2=pd.read_excel(root)
    #appending 
    col_T=df2.columns.get_loc("TIMESTAMP")
    # print(col_T)
    for num in range(0,len(df_data)):
        col=df2.columns.get_loc(df_data["label"][num].upper())
        # print(df_data["label"][row_])
        # print(df_data["time_"][row_])
        xf1 = openpyxl.load_workbook(filename=root)
        sheet1 = xf1['Sheet']
        sheet1.cell(row=ro,column=1,value=df_data["time_"][num])
        xf1.save(root)
        sheet1.cell(row=ro,column=col+1,value=df_data["value"][num])
        xf1.save(root)



def find_indx(len_sql_time,indx_list):
    indx_list.append(len_sql_time)    
    try:
        idx=int(reduce(lambda x, y:x+y, indx_list)) 
    except:
        idx=len_sql_time
    return idx

def recur_functn(veh_,sql_df,veh_df,veh_nam,curr_date,prev_date,first_timstmp,first_col,indx,indx_list,ro,strt_time):    
    # print(veh_name_list)
    for row_ in range(indx,len(sql_df)):
        # try:
        if indx<len(sql_df):
            if sql_df['time_'][row_]==first_timstmp:
                print(sql_df['time_'][row_])
                db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data` WHERE `date_` LIKE %(p)s and `time_` LIKE %(p1)s"
                sql_df_time,veh_df_time,veh_logr_time,veh_nam=create_connectn(curr_date,prev_date,db_qry,sql_df['time_'][row_])
                # db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data_testing` WHERE `date_` LIKE %(p)s and `time_` LIKE %(p1)s"
                sql_df_time,veh_df_time,veh_logr_time,veh_nam=create_connectn(prev_date,prev_date,db_qry,sql_df['time_'][row_])
                # print(veh_nam,veh_df_time)
                len_sql_time=len(sql_df_time)
                indx=find_indx(len_sql_time,indx_list)
                print("----------------------------------------------------------------------------------------------------------------------")
                print("Writing rows from index:",indx)
                print("----------------------------------------------------------------------------------------------------------------------")
                # print(sql_df_time['time_'])                  
                write_to_csv(sql_df_time,ro,veh_,veh_nam)
                ro+=1
                try:   
                    first_timstmp=sql_df['time_'][indx]  
                except:
                    break 
                
                return recur_functn(veh_,sql_df,veh_df,veh_nam,curr_date,prev_date,first_timstmp,first_col,indx,indx_list,ro,strt_time)
            else:
                pass
        else:
            break
        # except:
        #     pass
    time_tkn=time.time()-strt_time
    print("RRD created ,took seconds:",time_tkn)
    print('--------------------------------------------------------------------------------------------------------------------------------')
    # last_label_val(file_name,dir_path,prev_date)


def insert_data(veh_,sql_df,veh_df,veh_nam,curr_date,prev_date,len_veh_list,strt_time):
    print("----------------------------------------------------------------------------------------------------------------------")
    print(veh_)
    print("----------------------------------------------------------------------------------------------------------------------")
    first_timstmp=sql_df['time_'][0]
    first_col=sql_df['label'][0]
    count_col=0
    count_label=0
    ro=2
    indx_list=[]
    # res=search_vehicl(sql_df[vehicle_name],veh_,0,len(sql_df)-1)
    # if res!=-1:
    recur_data=recur_functn(veh_,sql_df,veh_df,veh_nam,curr_date,prev_date,first_timstmp,first_col,0,indx_list,ro,strt_time)


def write_RRD(prev_date,curr_date,veh_,len_veh_list):
    sys.setrecursionlimit(100000)
    print("Writing RRD")
    strt_time=time.time()
    # veh_nam=dict(zip(vehicle_df.viriciti_name,vehicle_df.fleet_name))
    # veh_logr=dict(zip(vehicle_df.viriciti_name,vehicle_df.logger))
    # # print(veh_logr)  
    # curr_date=datetime.now().strftime('%Y-%m-%d')
    # prev_date= (datetime.now() + timedelta(days=-1)).strftime('%Y-%m-%d')
    # db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data_testing` WHERE `date_` LIKE %(p)s"
    # sql_df,veh_df,veh_logr,veh_nam=create_connectn(curr_date,prev_date,db_qry,0)
    # count=0
    # for veh_ in veh_name_list:
    # try:
    db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data` WHERE `date_` LIKE %(p)s and `vehicle_name` LIKE %(p1)s"
    # db_qry="SELECT `vehicle_name`,`date_`,`time_`,`label`,`value` FROM `telematics`.`viriciti_raw_data_testing` WHERE `date_` LIKE %(p)s and `vehicle_name` LIKE %(p1)s"
    sql_df,veh_df,veh_logr,veh_nam,sql_df.columns=create_connectn(curr_date,prev_date,db_qry,veh_,0)
    len_veh_list=len(veh_name_list)
    # len_veh_list=1
    insert_data(veh_,sql_df,veh_df,veh_nam,curr_date,prev_date,len_veh_list,strt_time)
    
    # except:
    #     try:
    #         LogDirectory = os.path.join("E:\\","Company","truck_data","Viriciti","JOSN_files","logs", )
    #         LogFilepath = os.path.join(LogDirectory ,"Socket_{}.log".format(prev_date))            
    #         logging.basicConfig(filename=LogFilepath,filemode='a+',level=logging.CRITICAL)       
    #         exc_type, exc_value, exc_traceback = sys.exc_info()      
            
    #         logging.critical('{}:{} raised for message: {}'.format(exc_type,exc_value,msg))

    #     except:
    #         logging.exception("Exception occurred in create_RRD script")
        
    




# -----------------------------------------------------
# try:
#     sys.setrecursionlimit(100000)
#     write_RRD()
# except:
#     print("No data available for today!")
        
